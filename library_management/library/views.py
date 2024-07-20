from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from django.views.generic.list import ListView
from django.contrib import messages
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
import openpyxl


# List views

class AuthorListView(ListView):
    model = Author
    template_name = 'author_list.html'
    context_object_name = 'authors'


class BookListView(ListView):
    model = Book
    template_name = 'book_list.html'
    context_object_name = 'books'


class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = 'borrow_record_list.html'
    context_object_name = 'borrow_records'


# Create views

class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'add_author.html'
    success_url = reverse_lazy('author-list')


class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'add_book.html'
    success_url = reverse_lazy('book-list')


class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'add_borrow_record.html'
    success_url = reverse_lazy('borrowrecord-list')


def export_to_excel(request):
    # Create a workbook and sheets
    wb = openpyxl.Workbook()

    # Authors sheet
    ws1 = wb.active
    ws1.title = "Authors"
    ws1.append(['Name', 'Email', 'Bio'])
    for author in Author.objects.all():
        ws1.append([author.name, author.email, author.bio])

    # Books sheet
    ws2 = wb.create_sheet(title="Books")
    ws2.append(['Title', 'Genre', 'Published Date', 'Author'])
    for book in Book.objects.all():
        ws2.append([book.title, book.genre, book.published_date, book.author.name])

    # Borrow Records sheet
    ws3 = wb.create_sheet(title="Borrow Records")
    ws3.append(['User Name', 'Book', 'Borrow Date', 'Return Date'])
    for record in BorrowRecord.objects.all():
        ws3.append([record.user_name, record.book.title, record.borrow_date, record.return_date])

    # Save the workbook to a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
    wb.save(response)
    return response
